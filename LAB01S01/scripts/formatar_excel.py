import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

# Caminhos dos arquivos
csv_path = "../data/repositorios.csv"
excel_path = "../data/repositorios_formatado.xlsx"

# Verificar se o CSV existe
if not os.path.exists(csv_path):
    print(f"Erro: O arquivo '{csv_path}' não foi encontrado. Execute o script de coleta de dados primeiro.")
    exit()

# Carregar o CSV para um DataFrame do Pandas
df = pd.read_csv(csv_path)

# Salvar como Excel
df.to_excel(excel_path, index=False, engine="openpyxl")

# Abrir o arquivo Excel
wb = load_workbook(excel_path)
ws = wb.active

# Estilo para cabeçalhos
header_font = Font(bold=True, color="FFFFFF")  # Branco
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Azul
alignment = Alignment(horizontal="center", vertical="center")

# Aplicar estilos aos cabeçalhos
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment

# Ajustar a largura das colunas automaticamente
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Adicionar o símbolo de porcentagem nas colunas de Issues
for row in range(2, len(df) + 2):
    for col_num, column in enumerate(df.columns, start=1):
        if column in ["Issues Fechadas", "Issues Abertas"]:
            cell = ws.cell(row=row, column=col_num)
            # Adiciona o símbolo de porcentagem
            cell.value = f"{cell.value}%"
            cell.number_format = '0%'

# Salvar as alterações no Excel
wb.save(excel_path)

print(f"Planilha formatada e salva em '{excel_path}'")
