import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os

# Caminhos dos arquivos
csv_path = "../data/repositorios.csv"
excel_path = "../data/repositorios_formatado.xlsx"

# Verificar se o CSV existe
if not os.path.exists(csv_path):
    print(f"Erro: O arquivo '{csv_path}' não foi encontrado. Execute o script de coleta de dados primeiro.")
    exit()

# Carregar o CSV para um DataFrame do Pandas
df = pd.read_csv(csv_path)

# Converter colunas de data para o formato datetime e remover fuso horário
df["Data de Criação"] = pd.to_datetime(df["Data de Criação"]).dt.tz_localize(None)
df["Última Atualização"] = pd.to_datetime(df["Última Atualização"]).dt.tz_localize(None)

# Calcular a diferença em dias entre a data de criação e a última atualização
df["Dias de Atividade"] = (df["Última Atualização"] - df["Data de Criação"]).dt.days

# Salvar os dados atualizados como Excel
df.to_excel(excel_path, index=False, engine="openpyxl")

# Abrir o arquivo Excel
wb = load_workbook(excel_path)
ws = wb.active

# Alterar nome da aba principal
ws.title = "Listagem dos Repositórios"

# Função para ajustar a largura das colunas automaticamente
def ajustar_largura_colunas(worksheet):
    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        worksheet.column_dimensions[col_letter].width = max_length

# Estilo para cabeçalhos
header_font = Font(bold=True, color="FFFFFF")  # Texto branco
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Fundo azul
alignment = Alignment(horizontal="center", vertical="center")

# Aplicar estilos aos cabeçalhos
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment

# Ajustar largura das colunas na aba principal
ajustar_largura_colunas(ws)

# Criar uma aba separada para análise de linguagens
ws_lang = wb.create_sheet(title="Análise de Linguagens")

# Contar a quantidade de repositórios por linguagem
language_counts = df["Linguagem Primária"].value_counts()

# Adicionar os dados na nova aba
ws_lang.append(["Linguagem", "Quantidade"])
for lang, count in language_counts.items():
    ws_lang.append([lang, count])

# Ajustar largura das colunas na aba de linguagens
ajustar_largura_colunas(ws_lang)

# Criar Gráfico de Barras para Linguagens Populares
chart_lang = BarChart()
chart_lang.title = "Linguagens Mais Usadas nos Repositórios Populares"
chart_lang.x_axis.title = "Quantidade de Repositórios"
chart_lang.y_axis.title = "Linguagens"
chart_lang.type = "bar"  # Gráfico de barras horizontais
chart_lang.width = 21
chart_lang.height = 22.1

data = Reference(ws_lang, min_col=2, min_row=1, max_row=len(language_counts) + 1)
categories = Reference(ws_lang, min_col=1, min_row=2, max_row=len(language_counts) + 1)

chart_lang.add_data(data, titles_from_data=True)
chart_lang.set_categories(categories)

ws_lang.add_chart(chart_lang, "D2")  # Posição do gráfico

# Criar Gráfico de Barras para comparar tempo de atividade dos repositórios
if "Dias de Atividade" in df.columns:
    ws_atividade = wb.create_sheet(title="Tempo de Atividade")

    df_sorted = df.sort_values(by="Dias de Atividade", ascending=False)

    ws_atividade.append(["Repositório", "Dias de Atividade"])
    for _, row in df_sorted.iterrows():
        ws_atividade.append([row["Nome"], row["Dias de Atividade"]])

    # Ajustar largura das colunas na aba de tempo de atividade
    ajustar_largura_colunas(ws_atividade)

    chart_atividade = BarChart()
    chart_atividade.title = "Tempo de Atividade dos Repositórios"
    chart_atividade.x_axis.title = "Repositórios"
    chart_atividade.y_axis.title = "Dias de Atividade"
    chart_atividade.type = "col"
    chart_atividade.width = 26
    chart_atividade.height = 12

    data = Reference(ws_atividade, min_col=2, min_row=1, max_row=len(df_sorted) + 1)
    categories = Reference(ws_atividade, min_col=1, min_row=2, max_row=len(df_sorted) + 1)

    chart_atividade.add_data(data, titles_from_data=True)
    chart_atividade.set_categories(categories)

    ws_atividade.add_chart(chart_atividade, "D2")

# Salvar as alterações no Excel
wb.save(excel_path)

print(f"✅ Planilha formatada e salva em '{excel_path}' com gráficos e colunas ajustadas automaticamente")
